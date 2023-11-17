import numpy as np
import openpyxl
import re

from tabulate import tabulate

from velocities import *
from atmos import *
from aircraft import Aircraft

def extract_float_from_string(s):
    # Define a regular expression pattern to match floating-point numbers
    # This pattern will match numbers with optional positive/negative signs, integer part, optional decimal point, and fractional part.
    pattern = r"[-+]?\d*\.\d+|\d+\.\d*|\d+"

    # Use the findall() function to extract all matching floats from the input string
    floats = re.findall(pattern, s)

    # Convert the extracted strings to actual float values and store them in a list
    float_values = [float(num) for num in floats]
    return float_values
    l = []
    for t in s.split():
        try:
            l.append(float(t))
        except ValueError:
            pass
    return l

if __name__ == '__main__':

    aircraft = Aircraft()

    #sol = openpyxl.load_workbook(r'C:\Users\hugoa\OneDrive - polymtl.ca\Ecole\A2023\AER8375\TPs-Perfo\solutions\AER8375_TP4A_sol.xlsx')['Sheet1']
    sol = openpyxl.load_workbook(
        r'C:\Users\Admin\OneDrive - polymtl.ca\Poly & School\POLY\2023 Automne\Performance\TPs\solutions\AER8375_TP4A_sol.xlsx')['Sheet1']
    cases = {}

    for row in sol.iter_rows(min_row=16, max_row=18, min_col=1, max_col=4, values_only=True):
        case = {}

        if isinstance(row[2], str):
            hp = extract_float_from_string(row[2])[0] # pressure altitude
        else:
            hp = float(row[2])

        p = pressure_from_alt(hp)

        dISA = extract_float_from_string(row[3])
        if len(dISA):
            dISA = dISA[0]
        else:
            dISA = 0

        temp = temp_from_alt(hp) + dISA

        weight = float(row[1])

        case['hp'] = hp
        case['T'] = temp
        case['weight'] = weight

        cases[row[0]] = case

    results = {}
    for id in cases:
        case = cases[id]

        hp = case['hp']
        T = case['T']
        weight = case['weight']
        RM = 'MTOFN'
        Flaps = 20
        Vwind=0
        grad_RW=0


        p = pressure_from_alt(hp)
        dISA = get_delta_ISA(hp,T)
        # OUTPUTS

        V1Min, V1Max, VR, V2, V_LO_OEI, V_LO_AEO, V_35_AEO = aircraft.takeoff_run_velocities(weight, hp, dISA)

        headers = ['V1Min', 'V1Max', 'VR', 'V2', 'V_LO_OEI', 'V_LO_AEO', 'V_35_AEO']
        results[id] = {'V1Min':V1Min, 'V1Max':V1Max, 'VR':VR, 'V2':V2, 'V_LO_OEI':V_LO_OEI, 'V_LO_AEO':V_LO_AEO, 'V_35_AEO':V_35_AEO}



    # headers = ['Cas', 'Cd', 'Cl', 'L/D', 'Cdp', 'Cdi', 'Cdcomp', 'Cdcntl', 'Cdwm', 'Pousée Totale', 'Trainée', 'AOA', 'Nz_sw', 'Phi_sw', 'Nz_buffet']
    #
    print(tabulate([[name, *inner.values()] for name, inner in results.items()],
                    headers = headers,
                    tablefmt="github",
                    floatfmt=(".1f",".1f", ".1f", ".1f", ".1f", ".1f", ".1f", ".1f")))



